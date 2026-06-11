import difflib
import hashlib
import io
import json
import os
import secrets
from datetime import datetime, timezone, timedelta
from typing import List

WIB = timezone(timedelta(hours=7))

import pandas as pd
from fastapi import Depends, FastAPI, Form, HTTPException, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, RedirectResponse, PlainTextResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.templating import Jinja2Templates

import database
import runner
import tasks
import ai_detector

app = FastAPI(title="HydroLab")
templates = Jinja2Templates(directory="templates")
_security = HTTPBasic()

# Credentials — override via env vars HYDROLAB_USER and HYDROLAB_PASS in start.bat
_DASH_USER = os.environ.get("HYDROLAB_USER", "instruktur")
_DASH_PASS = os.environ.get("HYDROLAB_PASS", "hydrolab2026")


def require_instructor(credentials: HTTPBasicCredentials = Depends(_security)):
    ok_user = secrets.compare_digest(credentials.username.encode(), _DASH_USER.encode())
    ok_pass = secrets.compare_digest(credentials.password.encode(), _DASH_PASS.encode())
    if not (ok_user and ok_pass):
        raise HTTPException(
            status_code=401,
            detail="Login instruktur diperlukan.",
            headers={"WWW-Authenticate": "Basic"},
        )


# ── WebSocket manager ────────────────────────────────────────────────────────

class DashboardManager:
    def __init__(self):
        self._connections: List[WebSocket] = []

    async def connect(self, ws: WebSocket):
        await ws.accept()
        self._connections.append(ws)

    def disconnect(self, ws: WebSocket):
        if ws in self._connections:
            self._connections.remove(ws)

    async def broadcast(self, payload: dict):
        msg = json.dumps(payload)
        dead = []
        for ws in self._connections:
            try:
                await ws.send_text(msg)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(ws)


manager = DashboardManager()


# ── Startup ──────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    database.init_db()
    runner.ensure_data()
    runner.ensure_data_single()


# ── Student: join ────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def join_page(request: Request):
    return templates.TemplateResponse(request, "join.html")


@app.post("/join")
async def join(nim: str = Form(...), name: str = Form(...)):
    sid = database.create_session(nim, name)
    return RedirectResponse(f"/select-package/{sid}", status_code=303)


# ── Student: package selection ───────────────────────────────────────────────

@app.get("/select-package/{session_id}", response_class=HTMLResponse)
async def select_package_page(request: Request, session_id: int):
    session = database.get_session(session_id)
    if not session:
        return RedirectResponse("/")
    return templates.TemplateResponse(request, "select_package.html", {
        "session": session,
        "packages": tasks.PACKAGES,
    })


@app.post("/select-package/{session_id}")
async def select_package(session_id: int, package_id: int = Form(...)):
    session = database.get_session(session_id)
    if not session:
        return RedirectResponse("/")
    if package_id not in tasks.PACKAGES:
        return RedirectResponse(f"/select-package/{session_id}", status_code=303)
    database.set_package(session_id, package_id)
    return RedirectResponse(f"/student/{session_id}", status_code=303)


# ── Student: coding page ─────────────────────────────────────────────────────

@app.get("/student/{session_id}", response_class=HTMLResponse)
async def student_page(request: Request, session_id: int):
    session = database.get_session(session_id)
    if not session:
        return RedirectResponse("/")
    if not session["package_id"]:
        return RedirectResponse(f"/select-package/{session_id}")

    pkg_id = session["package_id"]
    total = tasks.get_package_total(pkg_id)
    all_done = session["current_task"] > total
    task = tasks.get_task_from_package(pkg_id, session["current_task"])
    if not task:
        task = tasks.get_task_from_package(pkg_id, total)

    return templates.TemplateResponse(request, "student.html", {
        "session": session,
        "task": task,
        "total_tasks": total,
        "all_done": all_done,
        "package": tasks.PACKAGES[pkg_id],
    })


@app.post("/run/{session_id}/{task_no}")
async def run_code(session_id: int, task_no: int, request: Request):
    body = await request.json()
    code = body.get("code", "")
    session = database.get_session(session_id)
    pkg_id = session["package_id"] if session else None
    result = runner.run_code(code, package_id=pkg_id, task_no=task_no)
    database.save_run(session_id, task_no, code, result["stdout"], result["stderr"])
    await manager.broadcast({"type": "run", "session_id": session_id})
    return result


@app.post("/submit/{session_id}/{task_no}")
async def submit_task(session_id: int, task_no: int, request: Request):
    body = await request.json()
    code = body.get("code", "")
    explanation = body.get("explanation", "").strip()
    session = database.get_session(session_id)
    pkg_id = session["package_id"] if session else None
    result = runner.run_code(code, package_id=pkg_id, task_no=task_no)

    # Run AI detection
    ai_analysis = ai_detector.detect_ai_patterns(code)
    ai_score = ai_analysis["score"]

    database.save_submission(
        session_id, task_no, code, result["stdout"],
        explanation=explanation, ai_score=ai_score
    )

    total = tasks.get_package_total(pkg_id)
    next_task = task_no + 1

    if next_task <= total:
        database.update_current_task(session_id, next_task)
        await manager.broadcast({"type": "submit", "session_id": session_id})
        return {
            "next_task": next_task,
            "done": False,
            "ai_analysis": ai_analysis,
        }
    else:
        database.update_current_task(session_id, next_task)
        await manager.broadcast({"type": "done", "session_id": session_id})
        return {
            "next_task": None,
            "done": True,
            "ai_analysis": ai_analysis,
        }


# ── Exam completed ───────────────────────────────────────────────────────────

@app.get("/exam-completed/{session_id}", response_class=HTMLResponse)
async def exam_completed_page(request: Request, session_id: int):
    session = database.get_session(session_id)
    if not session:
        return RedirectResponse("/")

    pkg_id = session.get("package_id")
    total = tasks.get_package_total(pkg_id) if pkg_id else 0

    # Guard: only accessible once all tasks are submitted
    if pkg_id is None or session["current_task"] <= total:
        return RedirectResponse(f"/student/{session_id}")

    # Completion token: deterministic, verifiable by instructor
    token_raw = f"{session['nim']}|{session_id}|{session['started_at']}"
    digest = hashlib.sha256(token_raw.encode()).hexdigest().upper()
    token = f"{digest[0:4]}-{digest[4:8]}-{digest[8:12]}-{digest[12:16]}"

    pkg_name = tasks.PACKAGES[pkg_id]["name"] if pkg_id in tasks.PACKAGES else "—"
    submitted_count = total  # all tasks completed to reach this page

    return templates.TemplateResponse(request, "completed.html", {
        "session": session,
        "token": token,
        "pkg_name": pkg_name,
        "total_tasks": total,
        "submitted_count": submitted_count,
        "completed_at": datetime.now(WIB).strftime("%d %B %Y, %H:%M:%S"),
    })


@app.get("/download-all/{session_id}")
async def download_all_code(session_id: int):
    session = database.get_session(session_id)
    if not session:
        return RedirectResponse("/")

    pkg_id = session.get("package_id")
    total = tasks.get_package_total(pkg_id) if pkg_id else 0
    if pkg_id is None or session["current_task"] <= total:
        return RedirectResponse(f"/student/{session_id}")

    detail = database.get_session_detail(session_id)
    submissions = detail["submissions"]
    if not submissions:
        return PlainTextResponse("Tidak ada kode yang tersimpan untuk session ini.", status_code=404)

    parts = []
    for sub in submissions:
        parts.append(f"=== Soal {sub['task_no']} ===")
        parts.append(sub["code"] or "")
        parts.append("")

    content = "\n".join(parts).strip() + "\n"
    filename = f"hydrolab_session{session_id}_all_tasks.txt"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return PlainTextResponse(content, headers=headers, media_type="text/plain")


# ── SEB quit endpoint ────────────────────────────────────────────────────────

@app.get("/seb-quit", response_class=HTMLResponse)
async def seb_quit():
    """Dedicated quit URL for Safe Exam Browser. Configure SEB Quit URL to this path."""
    html = """<!DOCTYPE html>
<html lang="id">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1"/>
  <title>Sesi Berakhir</title>
  <style>
    body { margin:0; display:flex; align-items:center; justify-content:center;
           min-height:100vh; background:#0d3b5e; font-family:'Segoe UI',sans-serif; color:#fff; text-align:center; }
    .msg { font-size:1.3rem; opacity:.85; }
    .icon { font-size:3.5rem; margin-bottom:16px; }
  </style>
</head>
<body>
  <div>
    <div class="icon">🔒</div>
    <div class="msg">Sesi ujian telah berakhir.<br>Safe Exam Browser sedang menutup…</div>
  </div>
</body>
</html>"""
    return HTMLResponse(content=html)


# ── Instructor dashboard ──────────────────────────────────────────────────────

@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(request: Request, _=Depends(require_instructor)):
    return templates.TemplateResponse(request, "dashboard.html", {
        "packages": tasks.PACKAGES,
    })


@app.get("/api/sessions")
async def api_sessions(_=Depends(require_instructor)):
    return database.get_all_sessions()


@app.get("/api/session/{session_id}")
async def api_session_detail(session_id: int, _=Depends(require_instructor)):
    data = database.get_session_detail(session_id)
    # Attach rubric for each submitted task
    pkg_id = data["session"].get("package_id")
    if pkg_id:
        for sub in data["submissions"]:
            t = tasks.get_task_from_package(pkg_id, sub["task_no"])
            sub["rubric"] = t["rubric"] if t else []
            sub["task_title"] = t["title"] if t else ""
    return data


@app.post("/api/clean-db")
async def clean_db(_=Depends(require_instructor)):
    database.clean_all()
    await manager.broadcast({"type": "cleaned"})
    return {"status": "ok", "message": "Semua data sesi berhasil dihapus."}


# ── Code similarity helper ────────────────────────────────────────────────────

def _code_sim(a: str, b: str) -> int:
    a = (a or '').strip()
    b = (b or '').strip()
    if not a or not b:
        return 0
    return round(difflib.SequenceMatcher(None, a, b).ratio() * 100)


# ── Build grading Excel ───────────────────────────────────────────────────────

def _build_grading_excel(sessions_df: pd.DataFrame, subs_df: pd.DataFrame) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # ── Palette ───────────────────────────────────────────────────────────────
    F_HDR      = PatternFill("solid", fgColor="0D3B5E")
    F_HDR2     = PatternFill("solid", fgColor="1A6FA8")
    F_RUBRIK   = PatternFill("solid", fgColor="E8F4FD")
    F_AI_HI    = PatternFill("solid", fgColor="FF6B6B")
    F_AI_MID   = PatternFill("solid", fgColor="FFD93D")
    F_AI_LOW   = PatternFill("solid", fgColor="FFE5CC")
    F_SIM_HI   = PatternFill("solid", fgColor="C0392B")
    F_SIM_MID  = PatternFill("solid", fgColor="E67E22")
    F_SIM_LOW  = PatternFill("solid", fgColor="F39C12")
    F_ALT      = PatternFill("solid", fgColor="F5F8FA")
    F_NILAI    = PatternFill("solid", fgColor="1A6FA8")
    F_PENALTI  = PatternFill("solid", fgColor="C0392B")
    F_FINAL    = PatternFill("solid", fgColor="1E8449")

    fw_b  = Font(color="FFFFFF", bold=True)
    fb    = Font(bold=True)
    center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_l   = Alignment(wrap_text=True, vertical="top")
    top_c   = Alignment(wrap_text=True, vertical="center", horizontal="center")

    task_numbers = sorted(subs_df['task_no'].unique().tolist()) if not subs_df.empty else []

    # ── 0_PANDUAN ─────────────────────────────────────────────────────────────
    wp = wb.create_sheet("0_PANDUAN")

    wp["A1"] = "PANDUAN PENILAIAN HYDROLAB — UPLOAD SHEET Soal_X KE CLAUDE UNTUK PENILAIAN"
    wp["A1"].font = Font(bold=True, size=13, color="0D3B5E")
    wp.merge_cells("A1:G1")
    wp.row_dimensions[1].height = 28

    wp["A3"] = "KRITERIA PENILAIAN"
    wp["A3"].font = Font(bold=True, size=11, color="0D3B5E")

    crit_hdr = ["No", "Kriteria", "Bobot", "Yang Diperiksa Claude", "Kolom Referensi"]
    for ci, h in enumerate(crit_hdr, 1):
        c = wp.cell(row=4, column=ci, value=h)
        c.fill = F_HDR; c.font = fw_b; c.alignment = center
    wp.row_dimensions[4].height = 30

    crit_rows = [
        (1, "Originalitas — Tanpa AI",    "30%",
         "Pola komentar berlebihan, variabel terlalu deskriptif (var_for_storing_result), struktur boilerplate ChatGPT",
         "AI_Risk (0–100). ≥70=🚨 ≥50=⚠️"),
        (2, "Anti-Plagiat antar mahasiswa","25%",
         "Kode identik/sangat mirip antar mahasiswa. Perhatikan urutan operasi, nama variabel, dan logika yang sama persis",
         "Sim_Max%, sheet DUPLIKASI"),
        (3, "Konsep & Logika Benar",       "30%",
         "Pendekatan sesuai soal, fungsi/metode tepat, hasil output masuk akal secara hidrologis",
         "Kolom Output + rubrik di baris 2 tiap sheet"),
        (4, "Dapat Dijalankan",            "15%",
         "Output terisi dan tidak ada traceback/error. Kosong = tidak jalan",
         "Kolom Output"),
    ]
    for ri, row_data in enumerate(crit_rows, 5):
        for ci, val in enumerate(row_data, 1):
            c = wp.cell(row=ri, column=ci, value=val)
            c.alignment = top_l
            if ri % 2 == 0:
                c.fill = F_ALT
    wp.row_dimensions[5].height = 20
    wp.row_dimensions[6].height = 40
    wp.row_dimensions[7].height = 40
    wp.row_dimensions[8].height = 20

    wp["A10"] = "PROMPT YANG DISARANKAN UNTUK CLAUDE:"
    wp["A10"].font = Font(bold=True, size=10)
    wp.merge_cells("A10:G10")

    prompt_text = (
        'Nilai setiap submission di sheet ini berdasarkan 4 kriteria: '
        '(1) Originalitas/tanpa AI 30%, (2) Anti-plagiat 25%, (3) Konsep benar 30%, (4) Runnable 15%. '
        'Kolom AI_Risk sudah dihitung otomatis (0–100). Kolom Sim_Max% adalah kemiripan kode dengan mahasiswa lain. '
        'Untuk tiap mahasiswa berikan: skor per kriteria (0–100), total nilai (0–100), dan catatan singkat. '
        'Format output: tabel dengan kolom NIM | Nama | Skor_Orig | Skor_Plagiat | Skor_Konsep | Skor_Run | Total | Catatan'
    )
    wp["A11"] = prompt_text
    wp["A11"].font = Font(italic=True, size=9, color="2F527A")
    wp["A11"].alignment = Alignment(wrap_text=True, vertical="top")
    wp.merge_cells("A11:G11")
    wp.row_dimensions[11].height = 70

    row_cur = 13
    wp.cell(row=row_cur, column=1, value="RUBRIK PER SOAL").font = Font(bold=True, size=11, color="0D3B5E")
    row_cur += 1

    for task_no in task_numbers:
        task_subs = subs_df[subs_df['task_no'] == task_no]
        pkg_id    = int(task_subs.iloc[0]['package_id']) if not task_subs.empty else None
        task_def  = tasks.get_task_from_package(pkg_id, task_no) if pkg_id else None
        title     = task_def['title'] if task_def else f"Soal {task_no}"

        c = wp.cell(row=row_cur, column=1, value=f"  Soal {task_no} — {title}")
        c.font = Font(bold=True, color="0D3B5E"); c.fill = F_RUBRIK
        wp.merge_cells(f"A{row_cur}:G{row_cur}")
        wp.row_dimensions[row_cur].height = 22
        row_cur += 1

        if task_def and task_def.get('rubric'):
            for ci, h in enumerate(["No", "Kriteria", "Poin Maks", "Catatan untuk AI Grader"], 1):
                c2 = wp.cell(row=row_cur, column=ci, value=h)
                c2.fill = F_HDR2; c2.font = fw_b; c2.alignment = center
            wp.row_dimensions[row_cur].height = 28
            row_cur += 1

            for r in task_def['rubric']:
                wp.cell(row=row_cur, column=1, value=r['no'])
                wp.cell(row=row_cur, column=2, value=r['kriteria'])
                wp.cell(row=row_cur, column=3, value=r['poin'])
                wp.cell(row=row_cur, column=4, value="Nilai penuh jika sempurna; setengah jika sebagian; 0 jika tidak ada")
                for ci in range(1, 5):
                    wp.cell(row=row_cur, column=ci).alignment = top_l
                row_cur += 1

        row_cur += 1

    wp.column_dimensions['A'].width = 5
    wp.column_dimensions['B'].width = 42
    wp.column_dimensions['C'].width = 10
    wp.column_dimensions['D'].width = 60
    wp.column_dimensions['E'].width = 35
    wp.column_dimensions['F'].width = 10
    wp.column_dimensions['G'].width = 10

    # ── Per-task sheets ───────────────────────────────────────────────────────
    task_sim_pairs: dict = {}

    for task_no in task_numbers:
        task_subs = subs_df[subs_df['task_no'] == task_no].copy().sort_values('nim').reset_index(drop=True)
        if task_subs.empty:
            continue

        pkg_id   = int(task_subs.iloc[0]['package_id'])
        task_def = tasks.get_task_from_package(pkg_id, task_no)
        title    = task_def['title'] if task_def else f"Soal {task_no}"

        nims  = task_subs['nim'].tolist()
        names = task_subs['name'].tolist()
        codes = task_subs['code'].fillna('').tolist()

        # Pairwise similarity
        max_sim : dict = {n: 0  for n in nims}
        sim_with: dict = {n: '' for n in nims}
        pairs = []

        for i in range(len(nims)):
            for j in range(i + 1, len(nims)):
                s = _code_sim(codes[i], codes[j])
                if s > max_sim[nims[i]]:
                    max_sim[nims[i]] = s; sim_with[nims[i]] = nims[j]
                if s > max_sim[nims[j]]:
                    max_sim[nims[j]] = s; sim_with[nims[j]] = nims[i]
                if s >= 50:
                    pairs.append({
                        'Soal': task_no,
                        'NIM_A': nims[i], 'Nama_A': names[i],
                        'NIM_B': nims[j], 'Nama_B': names[j],
                        'Similarity_%': s,
                        'Status': ('🚨 IDENTIK' if s >= 90 else ('⚠️ SANGAT MIRIP' if s >= 70 else '🔍 MIRIP')),
                    })

        task_sim_pairs[task_no] = pairs
        task_subs['Sim_Max%']       = task_subs['nim'].map(max_sim)
        task_subs['Mirip_Dengan']   = task_subs['nim'].map(sim_with)

        ws = wb.create_sheet(f"Soal_{task_no}")

        # Row 1: title
        ws.cell(row=1, column=1, value=f"Soal {task_no} — {title}  ({len(task_subs)} mahasiswa)")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="0D3B5E")
        ws.merge_cells("A1:I1")
        ws.row_dimensions[1].height = 24

        # Row 2: rubric summary (untuk Claude baca inline)
        if task_def and task_def.get('rubric'):
            rubric_txt = "  |  ".join([f"{r['no']}. {r['kriteria']} ({r['poin']}p)" for r in task_def['rubric']])
            ws.cell(row=2, column=1, value=f"RUBRIK: {rubric_txt}")
            ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="1A6FA8")
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells("A2:I2")
            ws.row_dimensions[2].height = 45

        # Row 3: column headers
        hdrs = ["NIM", "Nama", "Kode", "Output", "Penjelasan Mahasiswa",
                "AI_Risk\n(0-100)", "Sim_Max%", "Mirip_Dengan_NIM", "submitted_at"]
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.fill = F_HDR; c.font = fw_b; c.alignment = center
        ws.row_dimensions[3].height = 36

        # Data rows
        for ri, row in task_subs.iterrows():
            excel_row = ri + 4
            ai   = int(row.get('ai_score') or 0)
            sim  = int(row.get('Sim_Max%') or 0)

            ws.cell(row=excel_row, column=1, value=row['nim'])
            ws.cell(row=excel_row, column=2, value=row['name'])
            ws.cell(row=excel_row, column=3, value=row.get('code', ''))
            ws.cell(row=excel_row, column=4, value=row.get('output', ''))
            ws.cell(row=excel_row, column=5, value=row.get('explanation', ''))
            ws.cell(row=excel_row, column=6, value=ai)
            ws.cell(row=excel_row, column=7, value=sim)
            ws.cell(row=excel_row, column=8, value=row.get('Mirip_Dengan', ''))
            ws.cell(row=excel_row, column=9, value=row.get('submitted_at', ''))

            ws.row_dimensions[excel_row].height = 90

            # AI color
            ai_c = ws.cell(row=excel_row, column=6)
            if ai >= 70:
                ai_c.fill = F_AI_HI;  ai_c.font = Font(bold=True, color="FFFFFF")
            elif ai >= 50:
                ai_c.fill = F_AI_MID; ai_c.font = fb
            elif ai >= 30:
                ai_c.fill = F_AI_LOW

            # Similarity color
            sim_c = ws.cell(row=excel_row, column=7)
            if sim >= 90:
                sim_c.fill = F_SIM_HI;  sim_c.font = Font(bold=True, color="FFFFFF")
            elif sim >= 70:
                sim_c.fill = F_SIM_MID; sim_c.font = Font(bold=True, color="FFFFFF")
            elif sim >= 50:
                sim_c.fill = F_SIM_LOW; sim_c.font = fb

            for ci in range(1, 10):
                c = ws.cell(row=excel_row, column=ci)
                c.alignment = top_l if ci in (3, 4, 5) else top_c
                if ci not in (6, 7) and excel_row % 2 == 0:
                    if not c.fill or c.fill.patternType == "none":
                        c.fill = F_ALT

        col_w = [18, 22, 55, 42, 36, 11, 11, 20, 18]
        for ci, w in enumerate(col_w, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "C4"

    # ── DUPLIKASI sheet ───────────────────────────────────────────────────────
    wd = wb.create_sheet("DUPLIKASI")

    wd["A1"] = "LAPORAN KEMIRIPAN KODE — DETEKSI PLAGIARISME"
    wd["A1"].font = Font(bold=True, size=12, color="8B0000")
    wd.merge_cells("A1:G1")
    wd.row_dimensions[1].height = 24

    wd["A2"] = "Threshold tampil: ≥50%. Warna merah ≥90% (identik), oranye ≥70% (sangat mirip), kuning ≥50% (mirip). Diurutkan dari paling mirip."
    wd["A2"].font = Font(italic=True, size=9)
    wd["A2"].alignment = Alignment(wrap_text=True)
    wd.merge_cells("A2:G2")
    wd.row_dimensions[2].height = 28

    dup_hdrs = ["Soal", "NIM_A", "Nama_A", "NIM_B", "Nama_B", "Similarity_%", "Status"]
    for ci, h in enumerate(dup_hdrs, 1):
        c = wd.cell(row=3, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor="8B0000"); c.font = fw_b; c.alignment = center
    wd.row_dimensions[3].height = 28

    all_pairs = sorted(
        [p for pl in task_sim_pairs.values() for p in pl],
        key=lambda x: (-x['Similarity_%'], x['Soal'])
    )

    if all_pairs:
        for ri, pair in enumerate(all_pairs, 4):
            sim = pair['Similarity_%']
            wd.cell(row=ri, column=1, value=pair['Soal'])
            wd.cell(row=ri, column=2, value=pair['NIM_A'])
            wd.cell(row=ri, column=3, value=pair['Nama_A'])
            wd.cell(row=ri, column=4, value=pair['NIM_B'])
            wd.cell(row=ri, column=5, value=pair['Nama_B'])
            wd.cell(row=ri, column=6, value=sim)
            wd.cell(row=ri, column=7, value=pair['Status'])

            fill = F_SIM_HI if sim >= 90 else (F_SIM_MID if sim >= 70 else F_SIM_LOW)
            fw   = Font(bold=True, color="FFFFFF") if sim >= 70 else fb
            for ci in range(1, 8):
                c = wd.cell(row=ri, column=ci)
                if ci in (6, 7):
                    c.fill = fill; c.font = fw
                elif ri % 2 == 0:
                    c.fill = F_ALT
                c.alignment = top_c
    else:
        wd.cell(row=4, column=1, value="Tidak ada pasangan kode dengan similarity ≥ 50% — bagus!")
        wd.cell(row=4, column=1).font = Font(italic=True, color="1E8449")
        wd.merge_cells("A4:G4")

    for ci, w in enumerate([8, 18, 28, 18, 28, 14, 18], 1):
        wd.column_dimensions[get_column_letter(ci)].width = w
    wd.freeze_panes = "A4"

    # ── PENILAIAN sheet ───────────────────────────────────────────────────────
    wn = wb.create_sheet("PENILAIAN")

    wn["A1"] = "LEMBAR PENILAIAN FINAL — Isi kolom Nilai_SoalX setelah review AI/manual"
    wn["A1"].font = Font(bold=True, size=12, color="0D3B5E")
    n_extra = 3 + len(task_numbers) * 2 + 5
    wn.merge_cells(f"A1:{get_column_letter(n_extra)}1")
    wn.row_dimensions[1].height = 24

    p_hdrs = ["NIM", "Nama", "Jml_Submit"]
    for tn in task_numbers:
        p_hdrs += [f"Nilai_Soal{tn}", f"AI_Risk_S{tn}"]
    p_hdrs += ["Nilai_Total", "Penalti_AI(0-10)", "Penalti_Plagiat(0-10)", "Nilai_Final", "Catatan"]

    for ci, h in enumerate(p_hdrs, 1):
        c = wn.cell(row=2, column=ci, value=h)
        if h.startswith("Nilai_Soal") or h == "Nilai_Total":
            c.fill = F_NILAI
        elif "Penalti" in h:
            c.fill = F_PENALTI
        elif h == "Nilai_Final":
            c.fill = F_FINAL
        else:
            c.fill = F_HDR
        c.font = fw_b; c.alignment = center
    wn.row_dimensions[2].height = 36

    if not sessions_df.empty:
        for ri, (_, sess) in enumerate(sessions_df.sort_values('nim').iterrows(), 3):
            nim = sess['nim']
            submitted = len(subs_df[subs_df['nim'] == nim]) if not subs_df.empty else 0

            wn.cell(row=ri, column=1, value=nim)
            wn.cell(row=ri, column=2, value=sess['name'])
            wn.cell(row=ri, column=3, value=submitted)

            col = 4
            for tn in task_numbers:
                sub = subs_df[(subs_df['nim'] == nim) & (subs_df['task_no'] == tn)]
                ai  = int(sub.iloc[0]['ai_score']) if not sub.empty and sub.iloc[0]['ai_score'] else 0

                wn.cell(row=ri, column=col,   value="")  # kosong untuk diisi
                ai_c = wn.cell(row=ri, column=col+1, value=ai)
                if ai >= 70:
                    ai_c.fill = F_AI_HI;  ai_c.font = Font(bold=True, color="FFFFFF")
                elif ai >= 50:
                    ai_c.fill = F_AI_MID; ai_c.font = fb
                elif ai >= 30:
                    ai_c.fill = F_AI_LOW
                ai_c.alignment = top_c
                col += 2

            # Nilai_Total, Penalti_AI, Penalti_Plagiat, Nilai_Final, Catatan — kosong
            for _ in range(5):
                wn.cell(row=ri, column=col, value="")
                col += 1

            wn.row_dimensions[ri].height = 20
            for ci2 in range(1, col):
                wn.cell(row=ri, column=ci2).alignment = top_c

    wn.column_dimensions['A'].width = 18
    wn.column_dimensions['B'].width = 28
    wn.column_dimensions['C'].width = 12
    col_i = 4
    for _ in task_numbers:
        wn.column_dimensions[get_column_letter(col_i)].width = 13
        wn.column_dimensions[get_column_letter(col_i+1)].width = 11
        col_i += 2
    for w in [13, 16, 17, 13, 38]:
        wn.column_dimensions[get_column_letter(col_i)].width = w
        col_i += 1
    wn.freeze_panes = "C3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Export endpoint ───────────────────────────────────────────────────────────

@app.get("/export")
async def export_excel(_=Depends(require_instructor)):
    import sqlite3
    conn = sqlite3.connect(database.DB_PATH)

    sessions_df = pd.read_sql("""
        SELECT s.nim, s.name, s.package_id, s.run_count, s.started_at
        FROM sessions s ORDER BY s.nim
    """, conn)

    subs_df = pd.read_sql("""
        SELECT s.nim, s.name, s.package_id,
               sub.task_no, sub.code, sub.output,
               sub.explanation, sub.ai_score, sub.submitted_at
        FROM submissions sub
        JOIN sessions s ON s.id = sub.session_id
        ORDER BY s.nim, sub.task_no
    """, conn)

    conn.close()

    buf = _build_grading_excel(sessions_df, subs_df)
    timestamp = datetime.now(WIB).strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=HydroLab_Grading_{timestamp}.xlsx"},
    )


@app.websocket("/ws/dashboard")
async def dashboard_ws(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)
