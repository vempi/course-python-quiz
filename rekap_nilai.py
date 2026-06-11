"""
rekap_nilai.py — Proses hasil export HydroLab menjadi rekap nilai final.

Usage:
    python rekap_nilai.py HydroLab_Hasil_20260526_1000.xlsx
    python rekap_nilai.py HydroLab_Hasil_20260526_1000.xlsx output_nilai.xlsx
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


# ── Risk helpers ─────────────────────────────────────────────────────────────

def risk_label(score):
    if pd.isna(score):
        return "—"
    s = int(score)
    if s >= 70:
        return f"SANGAT TINGGI ({s})"
    elif s >= 50:
        return f"TINGGI ({s})"
    elif s >= 30:
        return f"SEDANG ({s})"
    else:
        return f"Rendah ({s})"


def build_catatan_ai(row, ai_cols):
    flags = []
    for col in ai_cols:
        task_no = col.replace("AI_Soal_", "")
        score = row.get(col)
        if pd.notna(score) and int(score) >= 50:
            flags.append(f"Soal {task_no}: {risk_label(score)}")
    return " | ".join(flags) if flags else "—"


# ── Main ─────────────────────────────────────────────────────────────────────

def rekap_nilai(input_file: str, output_file: str = None) -> str:
    xl = pd.ExcelFile(input_file)

    rekap       = pd.read_excel(xl, "1_Rekap")
    submissions = pd.read_excel(xl, "2_Submission")

    # Normalise column names
    rekap.columns       = [c.strip() for c in rekap.columns]
    submissions.columns = [c.strip() for c in submissions.columns]

    # Detect AI score column (header contains 'AI')
    ai_raw_col = next((c for c in submissions.columns if "AI" in c.upper()), None)

    # ── Build per-student base ───────────────────────────────────────────────
    students = rekap[["nim", "name", "total_submit", "total_run"]].copy()
    students.rename(columns={"nim": "NIM", "name": "Nama",
                              "total_submit": "Total Submit",
                              "total_run": "Total Run"}, inplace=True)

    # ── AI score pivot (one column per task) ────────────────────────────────
    ai_cols = []
    if ai_raw_col and not submissions.empty:
        pivot = submissions.pivot_table(
            index="nim", columns="task_no",
            values=ai_raw_col, aggfunc="max"
        ).reset_index()
        pivot.rename(columns={"nim": "NIM"}, inplace=True)

        task_numbers = [c for c in pivot.columns if c != "NIM"]
        ai_cols = [f"AI_Soal_{t}" for t in task_numbers]
        pivot.columns = ["NIM"] + ai_cols

        students = students.merge(pivot, on="NIM", how="left")

    # ── Max AI & catatan ─────────────────────────────────────────────────────
    if ai_cols:
        students["Max AI Score"] = students[ai_cols].max(axis=1)
        students["Catatan AI Detection"] = students.apply(
            lambda r: build_catatan_ai(r, ai_cols), axis=1
        )
    else:
        students["Max AI Score"] = 0
        students["Catatan AI Detection"] = "—"

    # ── Scoring columns (instruktur isi manual) ──────────────────────────────
    n_tasks = len(ai_cols) if ai_cols else int(rekap["total_submit"].max() or 5)
    nilai_cols = [f"Nilai Soal {i}" for i in range(1, n_tasks + 1)]
    for col in nilai_cols:
        students[col] = ""
    students["Nilai Total"] = ""
    students["Catatan Instruktur"] = ""

    # ── Column order ─────────────────────────────────────────────────────────
    ai_display_cols = [f"AI_Soal_{i}" for i in range(1, n_tasks + 1)
                       if f"AI_Soal_{i}" in students.columns]

    ordered = (
        ["NIM", "Nama", "Total Submit", "Total Run"]
        + nilai_cols
        + ["Nilai Total"]
        + ai_display_cols
        + ["Max AI Score", "Catatan AI Detection", "Catatan Instruktur"]
    )
    ordered = [c for c in ordered if c in students.columns]
    result = students[ordered].copy()

    # Rename AI_Soal_X → AI Soal X for display
    result.rename(columns={c: c.replace("AI_Soal_", "AI Soal ") for c in result.columns},
                  inplace=True)

    # ── Write Excel ───────────────────────────────────────────────────────────
    if output_file is None:
        output_file = str(Path(input_file).stem) + "_REKAP_NILAI.xlsx"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Rekap_Nilai", index=False)

        ws = writer.sheets["Rekap_Nilai"]
        _format_sheet(ws, result.columns.tolist(), n_tasks)

    print(f"✅  Rekap nilai disimpan ke: {output_file}")
    print(f"    {len(result)} mahasiswa | {n_tasks} soal")
    return output_file


# ── Formatting ────────────────────────────────────────────────────────────────

def _format_sheet(ws, headers, n_tasks):
    BLUE_DARK  = "0D3B5E"
    BLUE_MID   = "1A6FA8"
    GREEN_SOFT = "D6EFD8"
    RED_SOFT   = "FFCCCC"
    YELLOW_SOFT= "FFF3CD"
    ORANGE_SOFT= "FFE5CC"

    header_fill  = PatternFill("solid", fgColor=BLUE_DARK)
    nilai_fill   = PatternFill("solid", fgColor=BLUE_MID)
    header_font  = Font(color="FFFFFF", bold=True)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap         = Alignment(wrap_text=True, vertical="top")

    # Header row
    for cell in ws[1]:
        h = cell.value or ""
        if "Nilai" in h or h in ("Total Submit", "Total Run"):
            cell.fill = PatternFill("solid", fgColor=BLUE_MID)
        else:
            cell.fill = header_fill
        cell.font      = header_font
        cell.alignment = center

    ws.row_dimensions[1].height = 36

    # Data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            h = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
            val = cell.value

            # AI score cells — colour by risk
            if h.startswith("AI Soal") and isinstance(val, (int, float)):
                if val >= 70:
                    cell.fill = PatternFill("solid", fgColor="FF6B6B")
                elif val >= 50:
                    cell.fill = PatternFill("solid", fgColor="FFD93D")
                elif val >= 30:
                    cell.fill = PatternFill("solid", fgColor="FFE5CC")

            # Max AI Score
            if h == "Max AI Score" and isinstance(val, (int, float)):
                if val >= 70:
                    cell.fill = PatternFill("solid", fgColor=RED_SOFT)
                elif val >= 50:
                    cell.fill = PatternFill("solid", fgColor=YELLOW_SOFT)

            # Catatan AI Detection — flag rows with issues
            if h == "Catatan AI Detection" and val and val != "—":
                cell.fill = PatternFill("solid", fgColor=YELLOW_SOFT)

            cell.alignment = wrap

    # Column widths
    width_map = {
        "NIM": 22, "Nama": 28,
        "Total Submit": 14, "Total Run": 12,
        "Nilai Total": 13, "Max AI Score": 14,
        "Catatan AI Detection": 45, "Catatan Instruktur": 35,
    }
    for i, h in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        if h in width_map:
            ws.column_dimensions[col_letter].width = width_map[h]
        elif h.startswith("Nilai Soal"):
            ws.column_dimensions[col_letter].width = 13
        elif h.startswith("AI Soal"):
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 14

    # Freeze header + NIM+Nama columns
    ws.freeze_panes = "C2"


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_f  = sys.argv[1]
    output_f = sys.argv[2] if len(sys.argv) > 2 else None
    rekap_nilai(input_f, output_f)
